from django.shortcuts import render
from django.db.models import Sum, Count, F, DecimalField, Avg, Q, Max
from django.db.models.functions import TruncMonth, ExtractMonth, ExtractYear
from django.utils import timezone
from datetime import timedelta
from books.models import Book
from orders.models import Order, OrderItem
from accounts.models import Users
import pandas as pd
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
import json
from django.core.serializers.json import DjangoJSONEncoder
from django.db.models import Sum, F, DecimalField, Count
from django.db.models.functions import TruncMonth
import openpyxl

def get_revenue_stats():
    """Thống kê doanh thu tổng + theo trạng thái + theo tháng và trạng thái + top sách bán chạy"""
    
    # Danh sách trạng thái
    STATUS_CHOICES = ['pending', 'processing', 'shipping', 'completed', 'cancelled']

    # Tính total_revenue cho từng trạng thái
    revenue_by_status = {}
    for status in STATUS_CHOICES:
        revenue = OrderItem.objects.filter(
            order__status=status
        ).aggregate(
            total=Sum(F('quantity') * F('price'), output_field=DecimalField())
        )['total'] or 0
        revenue_by_status[status] = revenue

    # Tổng doanh thu tất cả (không phân trạng thái)
    total_revenue = OrderItem.objects.aggregate(
        total_revenue=Sum(F('quantity') * F('price'), output_field=DecimalField())
    )['total_revenue'] or 0

    # Thống kê doanh thu theo tháng và trạng thái (lấy 6 tháng gần nhất)
    revenue_by_month_status = OrderItem.objects.annotate(
        month=TruncMonth('order__created_at')
    ).values('month', 'order__status').annotate(
        revenue=Sum(F('quantity') * F('price'), output_field=DecimalField()),
        order_count=Count('order__id', distinct=True)
    ).order_by('-month', 'order__status')[:6 * len(STATUS_CHOICES)]    # Lấy top 10 sách bán chạy (dựa trên số lượng bán)
    best_selling_books = OrderItem.objects.values(
        'book__title',  # Lấy tiêu đề sách
        'price'  # Lấy giá
    ).annotate(
        total_sold=Sum('quantity'),  # Tổng số lượng bán
        total_revenue=Sum(F('quantity') * F('price'))  # Tổng doanh thu
    ).exclude(
        book__title__isnull=True  # Loại bỏ các sách không có tiêu đề
    ).order_by('-total_sold')[:10]  # Sắp xếp giảm dần và lấy top 10

    return {
        'total_revenue': total_revenue,
        'revenue_by_status': revenue_by_status,
        'revenue_by_month_status': list(revenue_by_month_status),
        'best_selling_books': list(best_selling_books)  # Thêm top sách bán chạy
    }
    
def get_customer_order_stats():
    """Thống kê khách hàng và đơn hàng"""
    # Thống kê khách hàng
    total_customers = Users.objects.count()
    new_customers = Users.objects.filter(
        date_joined__gte=timezone.now() - timedelta(days=30)
    ).count()

    # Thống kê đơn hàng
    orders = Order.objects.all()
    total_orders = orders.count()
    orders_by_status = orders.values('status').annotate(count=Count('id'))

    return {
        'total_customers': total_customers,
        'new_customers': new_customers,
        'total_orders': total_orders,
        'orders_by_status': orders_by_status,
    }

def get_book_inventory_stats():
    """Thống kê tình trạng sách"""
    # Thống kê tổng quan về sách
    total_books = Book.objects.count()
    out_of_stock = Book.objects.filter(stock=0)
    low_stock = Book.objects.filter(stock__gt=0, stock__lte=5)

    return {
        'total_books': total_books,
        'out_of_stock_books': out_of_stock,
        'low_stock_books': low_stock
    }

def analytics_dashboard(request):
    """View chính cho dashboard analytics"""
    context = {
        **get_revenue_stats(),
        **get_customer_order_stats(),
        **get_book_inventory_stats()
    }
    return render(request, 'analytics/dashboard.html', context)

def book_inventory_detail(request):
    """Thống kê chi tiết về sách và kho"""
    # Xử lý tìm kiếm
    search_query = request.GET.get('search', '')
    sort_by = request.GET.get('sort', '-stock')  # Mặc định sắp xếp theo tồn kho giảm dần

    books = Book.objects.annotate(
        total_sold=Sum('orderitem__quantity')
    ).select_related('author', 'publisher')

    # Tìm kiếm
    if search_query:
        books = books.filter(
            Q(title__icontains=search_query) |
            Q(author__name__icontains=search_query) |
            Q(publisher__name__icontains=search_query)
        )

    # Sắp xếp
    books = books.order_by(sort_by)

    # Phân trang
    paginator = Paginator(books, 20)  # 20 items mỗi trang
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Xuất Excel
    # Phần xuất Excel trong book_inventory_detail
    if request.GET.get('export') == 'excel':
        data = []
        for book in books:
            status = (
                "Hết hàng" if book.stock == 0
                else "Sắp hết" if book.stock <= 5
                else "Còn hàng"
            )

            data.append({
                'Tên Sách': book.title,
                'Tác Giả': book.author.name if book.author else '',
                'NXB': book.publisher.name if book.publisher else '',
                'Tồn Kho': book.stock,
                'Đã Bán': book.total_sold or 0,
                'Giá ($)': book.price,
                'Trạng Thái': status
            })

        df = pd.DataFrame(data)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="book_inventory.xlsx"'
        df.to_excel(response, index=False, engine='openpyxl')
        return response

    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'sort_by': sort_by,
        'total_books': books.count(),
        'out_of_stock': books.filter(stock=0).count(),
        'low_stock': books.filter(stock__gt=0, stock__lte=5).count(),
    }

    return render(request, 'analytics/book_inventory_detail.html', context)

def customer_analysis_detail(request):
    """Chi tiết phân tích khách hàng và đơn hàng"""
    # Xử lý tìm kiếm
    search_query = request.GET.get('search', '')
    sort_by = request.GET.get('sort', '-total_spent')  # Mặc định sắp xếp theo tổng chi tiêu

    # Query khách hàng với tổng chi tiêu và số đơn hàng
    customers = Users.objects.annotate(
        total_spent=Sum(F('order__items__quantity') * F('order__items__price'),
                       output_field=DecimalField()),
        order_count=Count('order', distinct=True),
        last_order_date=Max('order__created_at')
    )

    # Tìm kiếm
    if search_query:
        customers = customers.filter(
            Q(username__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(phone__icontains=search_query)
        )

    # Sắp xếp
    customers = customers.order_by(sort_by)

    # Phân trang
    paginator = Paginator(customers, 20)  # 20 khách hàng mỗi trang
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Xuất Excel nếu được yêu cầu
    if request.GET.get('export') == 'excel':
        data = []
        for customer in customers:
            data.append({
                'Tên đăng nhập': customer.username,
                'Email': customer.email,
                'Số điện thoại': customer.phone or '',
                'Số đơn hàng': customer.order_count,
                'Tổng chi tiêu': customer.total_spent or 0,
                'Đơn hàng gần nhất': customer.last_order_date
            })

        df = pd.DataFrame(data)
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="customer_analysis.xlsx"'
        df.to_excel(response, index=False)
        return response

    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'sort_by': sort_by,
        'total_customers': customers.count(),
        'new_customers': customers.filter(
            date_joined__gte=timezone.now() - timedelta(days=30)
        ).count(),
        'active_customers': customers.filter(
            order__created_at__gte=timezone.now() - timedelta(days=30)
        ).distinct().count(),
    }

    return render(request, 'analytics/customer_analysis_detail.html', context)

@login_required
def customer_stats(request):
    """View thống kê chi tiết về khách hàng"""
    try:
        # Thống kê tổng quan
        total_customers = Users.objects.count()
        new_customers = Users.objects.filter(
            date_joined__gte=timezone.now() - timedelta(days=30)
        ).count()
        active_customers = Users.objects.filter(
            order__created_at__gte=timezone.now() - timedelta(days=30)
        ).distinct().count()

        # Tính trung bình chi tiêu của khách hàng
        avg_spent = Users.objects.annotate(
            total_spent=Sum(F('order__items__quantity') * F('order__items__price'),
                           output_field=DecimalField())
        ).aggregate(avg_spent=Avg('total_spent'))['avg_spent'] or 0

        # Top khách hàng theo chi tiêu
        top_customers = Users.objects.annotate(
            total_spent=Sum(F('order__items__quantity') * F('order__items__price'),
                           output_field=DecimalField()),
            order_count=Count('order', distinct=True)
        ).filter(total_spent__isnull=False).order_by('-total_spent')[:10]

        # Dữ liệu cho biểu đồ tăng trưởng - trả về QuerySet trực tiếp
        customer_growth = Users.objects.annotate(
            month=TruncMonth('date_joined')
        ).values('month').annotate(
            new_users=Count('id')
        ).order_by('-month')[:12]

        # Phân bố khách hàng theo chi tiêu - trả về QuerySet trực tiếp
        spending_distribution = Users.objects.annotate(
            total_spent=Sum(F('order__items__quantity') * F('order__items__price'),
                           output_field=DecimalField())
        ).exclude(total_spent__isnull=True).values('total_spent').annotate(
            count=Count('id')
        ).order_by('total_spent')

        # Trạng thái khách hàng
        now = timezone.now()
        active_users = Users.objects.filter(
            order__created_at__gte=now - timedelta(days=30)
        ).distinct().count()
        
        # Khách hàng có đơn hàng nhưng không hoạt động trong 30 ngày
        inactive_users = Users.objects.filter(
            order__isnull=False
        ).exclude(
            order__created_at__gte=now - timedelta(days=30)
        ).distinct().count()
        
        # Khách hàng chưa có đơn hàng nào
        no_orders = Users.objects.filter(order__isnull=True).count()

        context = {
            'total_customers': total_customers,
            'new_customers': new_customers,
            'active_customers': active_customers,
            'avg_spent': avg_spent,
            'top_customers': top_customers,
            'customer_growth': customer_growth,
            'spending_distribution': spending_distribution,
            'customer_status': {
                'active': active_users,
                'inactive': inactive_users,
                'no_orders': no_orders
            }
        }
        
        return render(request, 'analytics/customer_stats.html', context)

    except Exception as e:
        print(f"Error in customer_stats view: {str(e)}")
        # Trả về dữ liệu mặc định nếu có lỗi
        context = {
            'total_customers': 0,
            'new_customers': 0,
            'active_customers': 0,
            'avg_spent': 0,
            'top_customers': [],
            'customer_growth': [],
            'spending_distribution': [],
            'customer_status': {
                'active': 0,
                'inactive': 0,
                'no_orders': 0
            }
        }
        return render(request, 'analytics/customer_stats.html', context)

@login_required
def order_stats(request):
    """View thống kê chi tiết về đơn hàng"""
    # Thống kê tổng quan
    total_orders = Order.objects.count()
    new_orders = Order.objects.filter(
        created_at__gte=timezone.now() - timedelta(days=7)
    ).count()

    # Tổng doanh thu và giá trị trung bình đơn hàng
    revenue_data = OrderItem.objects.aggregate(
        total_revenue=Sum(F('quantity') * F('price'), output_field=DecimalField()),
        avg_order_value=Avg(F('quantity') * F('price'), output_field=DecimalField())
    )
    total_revenue = revenue_data['total_revenue'] or 0
    avg_order_value = revenue_data['avg_order_value'] or 0
    
    # Doanh thu theo tháng
    revenue_by_month = OrderItem.objects.annotate(
        month=TruncMonth('order__created_at')
    ).values('month').annotate(
        revenue=Sum(F('quantity') * F('price'), output_field=DecimalField())
    ).order_by('-month')[:12]

    # Convert month to string for JSON serialization
    # Giữ nguyên datetime object để template có thể dùng filter date
    revenue_by_month = [
        {
            'month': item['month'],
            'revenue': float(item['revenue']) if item['revenue'] else 0
        } for item in revenue_by_month
    ]

    # Trạng thái đơn hàng
    order_status = Order.objects.values('status').annotate(
        count=Count('id')
    ).order_by('status')

    # Chuyển đổi trạng thái sang tiếng Việt (nếu có STATUS_CHOICES)
    try:
        for status in order_status:
            status['status_display'] = dict(Order.STATUS_CHOICES)[status['status']]
    except:
        # Nếu không có STATUS_CHOICES, giữ nguyên
        for status in order_status:
            status['status_display'] = status['status']

    # Đơn hàng gần đây
    # Đơn hàng gần đây có phân trang
    recent_orders_qs = Order.objects.select_related('customer').order_by('-created_at')
    paginator = Paginator(recent_orders_qs, 10)  # 10 đơn hàng mỗi trang

    page_number = request.GET.get('page')
    recent_orders_page = paginator.get_page(page_number)
    # Phương thức thanh toán
    payment_methods = Order.objects.values('payment_method').annotate(
        count=Count('id')
    ).order_by('-count')

    context = {
        'total_orders': total_orders,
        'new_orders': new_orders,
        'total_revenue': total_revenue,
        'avg_order_value': avg_order_value,
        'revenue_by_month': revenue_by_month,
        'order_status': list(order_status),
        'recent_orders': recent_orders_page,
        'payment_methods': list(payment_methods)
    }

    return render(request, 'analytics/order_stats.html', context)

def export_books_to_excel(books):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    # Tiêu đề
    headers = ["Tên Sách", "Tác Giả", "NXB", "Tồn Kho", "Đã Bán", "Giá", "Trạng Thái"]
    ws.append(headers)

    for book in books:
        if book.stock == 0:
            status = "Hết hàng"
        elif book.stock <= 5:
            status = "Sắp hết"
        else:
            status = "Còn hàng"

        row = [
            book.title,
            book.author.name if book.author else '',
            book.publisher.name if book.publisher else '',
            book.stock,
            book.total_sold or 0,
            book.price,
            status,
        ]
        ws.append(row)

    # Tự động giãn cột
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # Tạo phản hồi
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="book_inventory.xlsx"'
    wb.save(response)
    return response
